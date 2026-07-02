"""
routes/attendance.py — Attendance endpoints.

Handles face recognition on frames, attendance queries, stats, and Excel export.
"""

import io
from datetime import datetime, date
from fastapi import APIRouter, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from typing import Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from database import (
    get_attendance,
    get_today_attendance,
    get_stats,
    get_person_attendance,
)
from face_engine import process_frame

router = APIRouter(prefix="/api/attendance", tags=["Attendance"])


@router.post("/recognize")
async def recognize_faces(frame: UploadFile = File(...)):
    """
    Accept a webcam frame (JPEG) and run face recognition.
    Returns detected faces with bounding boxes and match info.
    """
    image_bytes = await frame.read()
    result = await process_frame(image_bytes)
    return result


@router.get("/")
async def list_attendance(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None, description="Department / Class filter"),
):
    """Get attendance records, optionally filtered by date range and department."""
    records = await get_attendance(date_from=date_from, date_to=date_to, department=department)
    # Serialize datetime fields
    for r in records:
        if r.get("timestamp"):
            r["timestamp"] = r["timestamp"].isoformat()
    return records


@router.get("/today")
async def today_attendance():
    """Get today's attendance records."""
    records = await get_today_attendance()
    for r in records:
        if r.get("timestamp"):
            r["timestamp"] = r["timestamp"].isoformat()
    return records


@router.get("/stats")
async def attendance_stats():
    """Get dashboard statistics."""
    return await get_stats()


@router.get("/person/{person_id}")
async def person_attendance_history(person_id: str):
    """Get all attendance records for a specific person."""
    records = await get_person_attendance(person_id)
    for r in records:
        if r.get("timestamp"):
            r["timestamp"] = r["timestamp"].isoformat()
    return records


@router.get("/export")
async def export_excel(
    date_from: Optional[str] = Query(None, description="Start date (YYYY-MM-DD)"),
    date_to: Optional[str] = Query(None, description="End date (YYYY-MM-DD)"),
    department: Optional[str] = Query(None, description="Department / Class filter"),
):
    """
    Generate and download an Excel attendance report.
    Defaults to current month if no date range is specified.
    """
    # Default to current month
    if not date_from:
        today = date.today()
        date_from = today.replace(day=1).isoformat()
    if not date_to:
        date_to = date.today().isoformat()

    records = await get_attendance(date_from=date_from, date_to=date_to, department=department)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1a1a3e", end_color="1a1a3e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = ["Sr No", "Name", "Department / Class", "Date", "Check-In", "Check-Out", "Work Hours", "Status"]
    ws.append(headers)

    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for idx, record in enumerate(records, 1):
        row = [
            idx,
            record.get("name", "Unknown"),
            record.get("department", "General"),
            record.get("date", ""),
            record.get("check_in") or record.get("time") or "N/A",
            record.get("check_out") or "N/A",
            record.get("work_hours") or "N/A",
            record.get("status", "Present"),
        ]
        ws.append(row)

        # Style data rows
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=idx + 1, column=col_idx)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # Auto-fit column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_length + 4, 12)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    dept_suffix = f"_{department}" if department and department != "All" else ""
    filename = f"Attendance_Report_{date_from}_to_{date_to}{dept_suffix}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
